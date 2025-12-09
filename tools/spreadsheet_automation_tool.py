

import openpyxl
import os
import json
import logging
import pandas as pd
from typing import Union, List, Dict, Any, Optional

from tools.base_tool import BaseTool

logger = logging.getLogger(__name__)

class SpreadsheetAutomationTool(BaseTool):
    """
    A tool for automating spreadsheet operations, including reading/writing cells
    and ranges, creating/deleting sheets, and converting between Excel and CSV formats.
    """

    def __init__(self, tool_name: str = "SpreadsheetAutomation", data_dir: str = ".", **kwargs):
        super().__init__(tool_name=tool_name, **kwargs)
        self.data_dir = data_dir

    @property
    def description(self) -> str:
        return "Automates spreadsheet operations: read/write cells/ranges, create/delete sheets, convert formats."

    @property
    def parameters(self) -> Dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "operation": {"type": "string", "enum": ["read_sheet", "write_cell", "create_sheet", "delete_sheet", "list_sheets", "read_range", "write_range", "convert_to_csv", "convert_from_csv"]},
                "file_path": {"type": "string", "description": "Absolute path to the Excel file."},
                "sheet_name": {"type": "string", "description": "The name of the sheet."},
                "cell": {"type": "string", "description": "The cell address (e.g., 'A1')."},
                "value": {"type": ["string", "number", "boolean"], "description": "The value to write to the cell."},
                "cell_range": {"type": "string", "description": "The cell range (e.g., 'A1:B5')."},
                "start_cell": {"type": "string", "description": "The starting cell for writing a range."},
                "data": {"type": "array", "items": {"type": "array", "items": {"type": ["string", "number", "boolean", "null"]}}, "description": "Data to write as a list of lists."},
                "csv_file_path": {"type": "string", "description": "Absolute path to the CSV file."}
            },
            "required": ["operation", "file_path"]
        }

    def _load_workbook(self, file_path: str, create_if_not_exists: bool = False):
        """Helper to load workbook."""
        if not os.path.isabs(file_path): raise ValueError(f"File path must be absolute: '{file_path}'")
        if os.path.exists(file_path):
            return openpyxl.load_workbook(file_path)
        elif create_if_not_exists:
            workbook = openpyxl.Workbook()
            workbook.save(file_path)
            return workbook
        else:
            raise FileNotFoundError(f"Workbook not found at {file_path}")

    def read_sheet(self, file_path: str, sheet_name: Optional[str] = None) -> List[List[Any]]:
        """Reads data from a specific sheet."""
        workbook = self._load_workbook(file_path)
        if sheet_name: sheet = workbook[sheet_name]
        else: sheet = workbook.active

        data = []
        for row in sheet.iter_rows(values_only=True): data.append(list(row))
        return data

    def write_cell(self, file_path: str, sheet_name: str, cell: str, value: Any) -> Dict[str, Any]:
        """Writes content to a specific cell."""
        workbook = self._load_workbook(file_path, create_if_not_exists=True)
        if sheet_name in workbook.sheetnames: sheet = workbook[sheet_name]
        else: sheet = workbook.create_sheet(title=sheet_name)
        
        sheet[cell] = value
        workbook.save(file_path)
        return {"status": "success", "message": f"Wrote '{value}' to cell '{cell}' in sheet '{sheet_name}' of '{os.path.basename(file_path)}'."}

    def create_sheet(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """Creates a new sheet."""
        workbook = self._load_workbook(file_path, create_if_not_exists=True)
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(title=sheet_name)
            workbook.save(file_path)
            return {"status": "success", "message": f"Created new sheet '{sheet_name}' in '{os.path.basename(file_path)}'."}
        else:
            raise ValueError(f"Sheet '{sheet_name}' already exists.")

    def delete_sheet(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """Deletes a sheet."""
        workbook = self._load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
            workbook.save(file_path)
            return {"status": "success", "message": f"Deleted sheet '{sheet_name}' from '{os.path.basename(file_path)}'."}
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found.")

    def list_sheets(self, file_path: str) -> List[str]:
        """Lists all sheet names in a workbook."""
        workbook = self._load_workbook(file_path)
        return workbook.sheetnames

    def read_range(self, file_path: str, sheet_name: str, cell_range: str) -> List[List[Any]]:
        """Reads data from a specific cell range."""
        workbook = self._load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = []
        for row in sheet[cell_range]: data.append([cell.value for cell in row])
        return data

    def write_range(self, file_path: str, sheet_name: str, start_cell: str, data: List[List[Any]]) -> Dict[str, Any]:
        """Writes data to a specific cell range."""
        workbook = self._load_workbook(file_path, create_if_not_exists=True)
        if sheet_name in workbook.sheetnames: sheet = workbook[sheet_name]
        else: sheet = workbook.create_sheet(title=sheet_name)

        start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
        for r_idx, row_data in enumerate(data):
            for c_idx, value in enumerate(row_data):
                sheet.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)
        workbook.save(file_path)
        return {"status": "success", "message": f"Wrote data to range starting at '{start_cell}' in sheet '{sheet_name}' of '{os.path.basename(file_path)}'."}

    def convert_to_csv(self, file_path: str, csv_file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """Converts an Excel sheet to a CSV file."""
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        df.to_csv(csv_file_path, index=False)
        return {"status": "success", "message": f"Converted '{os.path.basename(file_path)}' to '{os.path.basename(csv_file_path)}'."}

    def convert_from_csv(self, csv_file_path: str, file_path: str, sheet_name: str = "Sheet1") -> Dict[str, Any]:
        """Converts a CSV file to an Excel sheet."""
        df = pd.read_csv(csv_file_path)
        df.to_excel(file_path, sheet_name=sheet_name, index=False)
        return {"status": "success", "message": f"Converted '{os.path.basename(csv_file_path)}' to '{os.path.basename(file_path)}'."}

    def execute(self, operation: str, file_path: str, **kwargs: Any) -> Any:
        try:
            if operation == "read_sheet":
                return self.read_sheet(file_path, kwargs.get("sheet_name"))
            elif operation == "write_cell":
                sheet_name = kwargs.get("sheet_name")
                cell = kwargs.get("cell")
                value = kwargs.get("value")
                if not all([sheet_name, cell, value is not None]):
                    raise ValueError("Missing 'sheet_name', 'cell', or 'value' for 'write_cell' operation.")
                return self.write_cell(file_path, sheet_name, cell, value)
            elif operation == "create_sheet":
                sheet_name = kwargs.get("sheet_name")
                if not sheet_name:
                    raise ValueError("Missing 'sheet_name' for 'create_sheet' operation.")
                return self.create_sheet(file_path, sheet_name)
            elif operation == "delete_sheet":
                sheet_name = kwargs.get("sheet_name")
                if not sheet_name:
                    raise ValueError("Missing 'sheet_name' for 'delete_sheet' operation.")
                return self.delete_sheet(file_path, sheet_name)
            elif operation == "list_sheets":
                return self.list_sheets(file_path)
            elif operation == "read_range":
                sheet_name = kwargs.get("sheet_name")
                cell_range = kwargs.get("cell_range")
                if not all([sheet_name, cell_range]):
                    raise ValueError("Missing 'sheet_name' or 'cell_range' for 'read_range' operation.")
                return self.read_range(file_path, sheet_name, cell_range)
            elif operation == "write_range":
                sheet_name = kwargs.get("sheet_name")
                start_cell = kwargs.get("start_cell")
                data = kwargs.get("data")
                if not all([sheet_name, start_cell, data]):
                    raise ValueError("Missing 'sheet_name', 'start_cell', or 'data' for 'write_range' operation.")
                return self.write_range(file_path, sheet_name, start_cell, data)
            elif operation == "convert_to_csv":
                csv_file_path = kwargs.get("csv_file_path")
                if not csv_file_path:
                    raise ValueError("Missing 'csv_file_path' for 'convert_to_csv' operation.")
                return self.convert_to_csv(file_path, csv_file_path, kwargs.get("sheet_name"))
            elif operation == "convert_from_csv":
                csv_file_path = kwargs.get("csv_file_path")
                if not csv_file_path:
                    raise ValueError("Missing 'csv_file_path' for 'convert_from_csv' operation.")
                return self.convert_from_csv(csv_file_path, file_path, kwargs.get("sheet_name"))
            else:
                raise ValueError(f"Invalid operation '{operation}'.")
        except Exception as e:
            self.logger.error(f"An error occurred during spreadsheet automation: {e}")
            return {"status": "error", "message": str(e)}

if __name__ == '__main__':
    print("Demonstrating SpreadsheetAutomationTool functionality...")
    temp_dir = "temp_spreadsheet_data"
    if os.path.exists(temp_dir): import shutil; shutil.rmtree(temp_dir)
    os.makedirs(temp_dir)
    
    spreadsheet_tool = SpreadsheetAutomationTool(data_dir=temp_dir)
    
    excel_file = os.path.join(temp_dir, "my_data.xlsx")
    csv_file = os.path.join(temp_dir, "output.csv")

    try:
        # 1. Create a new sheet and write a cell
        print("\n--- Creating sheet 'Sales' and writing to A1 ---")
        spreadsheet_tool.execute(operation="create_sheet", file_path=excel_file, sheet_name="Sales")
        spreadsheet_tool.execute(operation="write_cell", file_path=excel_file, sheet_name="Sales", cell="A1", value="Product")
        spreadsheet_tool.execute(operation="write_cell", file_path=excel_file, sheet_name="Sales", cell="B1", value="Revenue")
        print("Sheet created and cells written.")

        # 2. Write a range of data
        print("\n--- Writing data range to 'Sales' sheet ---")
        data_to_write = [
            ["Widget A", 1500],
            ["Gadget B", 2000],
            ["Thing C", 750]
        ]
        spreadsheet_tool.execute(operation="write_range", file_path=excel_file, sheet_name="Sales", start_cell="A2", data=data_to_write)
        print("Data range written.")

        # 3. Read the sheet
        print("\n--- Reading 'Sales' sheet ---")
        sheet_data = spreadsheet_tool.execute(operation="read_sheet", file_path=excel_file, sheet_name="Sales")
        print(json.dumps(sheet_data, indent=2))

        # 4. Convert to CSV
        print("\n--- Converting 'my_data.xlsx' to CSV ---")
        spreadsheet_tool.execute(operation="convert_to_csv", file_path=excel_file, csv_file_path=csv_file, sheet_name="Sales")
        print("Converted to CSV.")

        # 5. List sheets
        print("\n--- Listing sheets in 'my_data.xlsx' ---")
        sheets = spreadsheet_tool.execute(operation="list_sheets", file_path=excel_file)
        print(json.dumps(sheets, indent=2))

    except Exception as e:
        print(f"\nAn error occurred: {e}")
    finally:
        if os.path.exists(temp_dir): import shutil; shutil.rmtree(temp_dir)
        print(f"\nCleaned up temporary directory '{temp_dir}'.")
